"""Document conversion pipeline for OpenKB."""
from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass, field
from pathlib import Path

import pymupdf
from markitdown import MarkItDown
from openpyxl import load_workbook

from openkb.config import load_config
from openkb.images import copy_relative_images, extract_base64_images, convert_pdf_with_images
from openkb.state import HashRegistry

logger = logging.getLogger(__name__)


@dataclass
class ConvertResult:
    """Result returned by :func:`convert_document`."""

    raw_path: Path | None = None
    source_path: Path | None = None
    is_long_doc: bool = False
    skipped: bool = False
    file_hash: str | None = None  # For deferred hash registration


def get_pdf_page_count(path: Path) -> int:
    """Return the number of pages in the PDF at *path* using pymupdf."""
    with pymupdf.open(str(path)) as doc:
        return doc.page_count


def convert_xlsx_streaming(path: Path, *, max_rows: int = 5000, max_cols: int = 64) -> str:
    """Convert .xlsx to markdown using a memory-safe streaming reader.

    This avoids large in-memory DataFrames and pathological worksheet ranges
    that can cause extreme RAM usage in generic converters.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            lines.append("")
            rows_written = 0
            empty_streak = 0

            for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
                vals = []
                non_empty = False
                for cell in row:
                    if cell is None:
                        vals.append("")
                        continue
                    text = str(cell).strip()
                    vals.append(text)
                    if text:
                        non_empty = True

                if not non_empty:
                    empty_streak += 1
                    if rows_written == 0:
                        continue
                    # Stop after sustained empty tail to avoid scanning sparse sheets.
                    if empty_streak >= 200:
                        break
                else:
                    empty_streak = 0

                if non_empty:
                    lines.append(" | ".join(vals).rstrip())
                    rows_written += 1

            if rows_written == 0:
                lines.append("(No non-empty cells found within scan limits.)")

            lines.append("")
    finally:
        wb.close()

    return "\n".join(lines)


def convert_document(src: Path, kb_dir: Path) -> ConvertResult:
    """Convert a document and integrate it into the knowledge base.

    Steps:
    1. Hash-check — skip if already known.
    2. Copy source to ``raw/``.
    3. If PDF and page count >= threshold → return :attr:`ConvertResult.is_long_doc`.
    4. If ``.md`` — read, process relative images, save to ``wiki/sources/``.
    5. Otherwise — run MarkItDown, extract base64 images, save to ``wiki/sources/``.
    6. Register hash in the registry.
    """
    # ------------------------------------------------------------------
    # Load config & state
    # ------------------------------------------------------------------
    openkb_dir = kb_dir / ".openkb"
    config = load_config(openkb_dir / "config.yaml")
    threshold: int = config.get("pageindex_threshold", 20)
    registry = HashRegistry(openkb_dir / "hashes.json")

    # ------------------------------------------------------------------
    # 1. Hash check
    # ------------------------------------------------------------------
    file_hash = HashRegistry.hash_file(src)
    if registry.is_known(file_hash):
        logger.info("Skipping already-known file: %s", src.name)
        return ConvertResult(skipped=True)

    # ------------------------------------------------------------------
    # 2. Copy to raw/
    # ------------------------------------------------------------------
    raw_dir = kb_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    raw_dest = raw_dir / src.name
    if raw_dest.resolve() != src.resolve():
        shutil.copy2(src, raw_dest)

    # ------------------------------------------------------------------
    # 3. PDF long-doc detection
    # ------------------------------------------------------------------
    if src.suffix.lower() == ".pdf":
        page_count = get_pdf_page_count(src)
        if page_count >= threshold:
            logger.info(
                "Long PDF detected (%d pages >= %d threshold): %s",
                page_count,
                threshold,
                src.name,
            )
            return ConvertResult(raw_path=raw_dest, is_long_doc=True, file_hash=file_hash)

    # ------------------------------------------------------------------
    # 4/5. Convert to Markdown
    # ------------------------------------------------------------------
    sources_dir = kb_dir / "wiki" / "sources"
    sources_dir.mkdir(parents=True, exist_ok=True)
    images_dir = kb_dir / "wiki" / "sources" / "images" / src.stem
    images_dir.mkdir(parents=True, exist_ok=True)

    doc_name = src.stem

    if src.suffix.lower() == ".md":
        markdown = src.read_text(encoding="utf-8")
        markdown = copy_relative_images(markdown, src.parent, doc_name, images_dir)
    elif src.suffix.lower() == ".pdf":
        # Use pymupdf dict-mode for PDFs: text + images inline at correct positions
        markdown = convert_pdf_with_images(src, doc_name, images_dir)
    elif src.suffix.lower() == ".xlsx":
        markdown = convert_xlsx_streaming(src)
    else:
        # Non-PDF, non-MD: use markitdown (docx, pptx, html, etc.)
        mid = MarkItDown()
        result = mid.convert(str(src))
        markdown = result.text_content
        markdown = extract_base64_images(markdown, doc_name, images_dir)

    dest_md = sources_dir / f"{doc_name}.md"
    dest_md.write_text(markdown, encoding="utf-8")

    return ConvertResult(raw_path=raw_dest, source_path=dest_md, file_hash=file_hash)
